from flask import Flask, request , jsonify
import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook
import os
import pandas as pd
import qdrant_client
from dotenv import load_dotenv
import nest_asyncio
from llama_parse import LlamaParse
from llama_index.core import VectorStoreIndex, get_response_synthesizer
from llama_index.core.node_parser import SentenceSplitter
from llama_index.core.retrievers import VectorIndexRetriever
from llama_index.core.query_engine import RetrieverQueryEngine
from llama_index.core import Document
from llama_index.vector_stores.qdrant import QdrantVectorStore
from llama_index.embeddings.openai import OpenAIEmbedding
from llama_index.core.ingestion import IngestionPipeline
from llama_index.core.response.pprint_utils import pprint_response

from llama_index.core.query_engine import CustomQueryEngine
from llama_index.core.retrievers import BaseRetriever
from llama_index.core import get_response_synthesizer
from llama_index.core.response_synthesizers import BaseSynthesizer
from llama_index.llms.openai import OpenAI
from llama_index.core import PromptTemplate

app = Flask(__name__)

def extractor(file_path):
    # Load the workbook
    wb = load_workbook(filename=file_path, read_only=True)

    # Select the first sheet
    ws = wb.active

    # Initialize an empty list for the dictionaries
    data = []

    # Get the header row
    header = [cell for cell in next(ws.iter_rows(values_only=True))]

    # Create a dictionary mapping column names to indices
    col_index = {name: index for index, name in enumerate(header)}

    # Iterate over the rows of the worksheet
    for row in ws.iter_rows(min_row=2, values_only=True):
        # Skip rows that don't have values
        if not all(cell is None for cell in row):
            # Create a dictionary for the current row
            row_dict = {
                'title': row[col_index['Title']],
                'abstract': row[col_index['Abstract']],
                'description': row[col_index['English description']],
                'claims': row[col_index['Claims']]
            }

            # Add the dictionary to the list
            data.append(row_dict)

    return data

def newFileSaver(relevancy, file_path):

    # Load the workbook and select the active worksheet
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active

    # Define the headers
    relevancy_header = 'Relevancy predicted'
    comments_header = 'Comments made'

    empty_column = None
    for cell in sheet[1]:
        if cell.value is None:
            empty_column = cell.column
            break

    if empty_column == None:
        empty_column = sheet.max_column + 1
    
    # Add headers to the first row of the new columns
    sheet.cell(row=1, column=empty_column, value=relevancy_header)
    sheet.cell(row=1, column=empty_column+1, value=comments_header)

    # Add the values from the lists to the new columns
    for i, (relevancy, comment) in enumerate(relevancy, start=2):
        sheet.cell(row=i, column=empty_column, value=relevancy)
        sheet.cell(row=i, column=empty_column+1, value=comment)

    # Save the workbook
    workbook.save(filename=file_path)

    return file_path  


def extract_reason(text):
    parts = text.split("Reason: ", 1)
    if len(parts) > 1:
        return parts[1]
    else:
        return ""

def extract_related(text):
    return '1R1' in text

def backend(dict, user_query):

    # Load environment variables from the .env file
    load_dotenv()

    # Access the API key from the environment variables
    llama_api_key = os.getenv('LLAMA_CLOUD_API_KEY')

    if llama_api_key is None:
        raise ValueError("LLAMA_CLOUD_API_KEY not found in environment variables")

    documents = []
    for key, val in dict.items():
        temp = key + ": " + val
        documents.append(Document(text=temp))

    client = qdrant_client.QdrantClient(location=":memory:")
    vector_store = QdrantVectorStore(client=client, collection_name="test_store")

    pipeline = IngestionPipeline(
        transformations=[
            SentenceSplitter(chunk_size=128, chunk_overlap=5),
            OpenAIEmbedding(),
        ],
        vector_store=vector_store,
    )

    # Ingest directly into a vector db
    pipeline.run(documents=documents)
    #print(vector_store)
    index = VectorStoreIndex.from_vector_store(vector_store)

    
    class RAGStringQueryEngine(CustomQueryEngine):
        """RAG String Query Engine."""

        retriever: BaseRetriever
        response_synthesizer: BaseSynthesizer
        llm: OpenAI
        qa_prompt: PromptTemplate

        def custom_query(self, query_str: str):
            nodes = self.retriever.retrieve(query_str)

            context_str = "\n\n".join([n.node.get_content() for n in nodes])
            #print(context_str)
            #print('--'*50)
            response = self.llm.complete(
                qa_prompt.format(context_str=context_str, query_str=query_str)
            )

            return str(response)

    # configure retriever
    retriever1 = VectorIndexRetriever(
        index=index,
        similarity_top_k=5,
    )

    # configure response synthesizer
    response_synthesizer1 = get_response_synthesizer(
        response_mode="tree_summarize",
        # streaming= True,
    )

    qa_prompt = PromptTemplate(
        "You are an AI assistant that predicts relevancy of a 'Document' with a certain 'Statement'. If it's a little relevant then return output as '1R1', otherwise '0R0'. If output is '1R1', then state the 'Reason'  which makes it relevant with the help of information present in 'Document'. \n"
       "For example 1:\n"
             
       "Document:" + ''' title: Composition, application of the composition, cosmetic preparation hydrogel bio-mask in the form of a compress, method of manufacturing the preparation
       Background of the invention.
       hydrogel bio-mask composed of natural materials and active ingredients, designed for cosmetic applications to enhance skin health. The hydrogel matrix provides a natural and effective medium for delivering active ingredients to the skin. the composition of the hydrogel bio-mask and its natural active ingredients. The following are the key points regarding the specific ingredients mentioned
        Hydrogel Matrix: The document emphasizes the use of a hydrogel matrix obtained from natural sources. Natural Active Ingredients: The hydrogel bio-mask includes various natural active ingredients intended for cosmetic use.''' + "\n"
       "Statement:" + "Mannuronic acid or avocado or alginate from algae should be used in a cosmetic formulation for any skin claim" + "\n"
       "Output: '0R0' \n"
       "Reason:  It is not mentioning the use of Mannuronic acid, alginate, or avocado but having skin claim for cosmetics \n"
             
       "For example 2:\n"
             
       "Document:" + ''' the use of mannuronic acid derivatives and alginate from algae in cosmetic formulations aimed at improving skin health by providing anti-photoaging benefits, moisture retention, antioxidant protection, and enzyme inhibition. The derivatives form an invisible film on the skin, protecting against UV damage and maintaining a moist environment. They exhibit strong antioxidant capabilities and inhibit enzymes like tyrosinase and elastase, reducing melanin production and collagen degradation.
        The primary focus of the patent is on alginate oligosaccharide derivatives derived from brown algae. These are used for their moisture absorption, antioxidation, and enzyme inhibition properties in skincare products. ''' + "\n"
       "Statement:" + "Mannuronic acid or avocado or alginate from algae should be used in a cosmetic formulation for any skin claim" + "\n"
       "Output: '1R1' \n"
       "Reason :  Mannuronic acid and alginate from algae is used for different skin claims in a cosmetic product \n"
 
        "For example 3:\n"
 
        "Document:" + ''' title: Use of brown algae water extract for preparing blue light resistant skin external product
       Background of the invention.
       using brown algae extract containing fucoidan for preparing topical skin care products that protect against blue light exposure. These products aim to improve skin health by reducing wrinkles and enhancing brightness, particularly for individuals frequently exposed to blue light. The invention emphasizes the benefits of fucoidan in long-term skin care.
       The present invention provides a use of a brown algae extract for preparing a skin topical product for anti-blue light, wherein the product is provided to a subject exposed to blue light, and the brown algae extract contains fucoidan.''' + "\n"
       "Statement:" + "Mannuronic acid or avocado or alginate from algae should be used in a cosmetic formulation for any skin claim" + "\n"
       "Output: '1R1' \n"
       "Reason: Alginate from Brown Algae is used for protecting against blue light in skincare products \n"
 
       "Using the below given Document and Statement , provide the Output and Reason"
        "Document: {context_str}\n"
        "Statement: {query_str}\n"
        "Output: "
        "Reason: "
    )

    llm = OpenAI(model="gpt-3.5-turbo")

    query_engine = RAGStringQueryEngine(
        retriever=retriever1,
        response_synthesizer=response_synthesizer1,
        llm=llm,
        qa_prompt=qa_prompt,
    )
    # query_engine = index.as_query_engine(response_mode="tree_summarize",verbose=True,)
    response = query_engine.query(user_query)
    # print(response)
    pprint_response(response, show_source=True)
    response_str = str(response)
    
    related = extract_related(response_str)
    reason = extract_reason(response_str)
    # print(related, reason)
    return (related, reason)



@app.route("/", methods=['GET', 'POST'])
def process_file():

    data = request.json

    query = data.get('query')
    file_path = data.get('file_path')
    #print(file_path)
    datalist = extractor(file_path)

    if datalist[-1]['title'] == None:
        datalist.pop()
    
    #print(datalist)
    #backend(datalist[0])
    # Backend process begins

    # Initialize the relevancy list
    relevancy = []

    # Iterate over each dictionary in the datalist
    for dict_item in datalist:
        # Call the backend function with the current dictionary
        result = backend(dict_item, query)
        # Check the first element of the tuple and set 'R' or 'NR' accordingly
        status = "R" if result[0] else "NR"
        # Append the modified result to the relevancy list
        relevancy.append((status, result[1]))

    outputFilePath = newFileSaver(relevancy, file_path)
    
    return jsonify({'Path': outputFilePath})

if __name__ == "__main__":
    app.run(host='0.0.0.0', port=5000, debug=True)
